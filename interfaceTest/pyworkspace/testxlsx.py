import openpyxl

# wb = openpyxl.load_workbook('C:\\wq\\pyworkspace\\xlinkweb\\test_data\\login.xlsx')
# sheet = wb.get_sheet_by_name('locator')
# for i in range(2, sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         print sheet.cell(row=i,column=j).value
    # if sheet.cell(row=i,column=j).value == 'password_loc':
    #     print sheet.cell(row=i,column=2).value
    #     break
    # if sheet.cell(row=i,column=2) =='password_loc':
    #     print sheet.cell(row=i,column=i)
    #     break

#wb.close()

def get_locator_value(file_path, sheet_name, loc_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=1).value == loc_name:
            break
    wb.close()
    return ws.cell(row=i, column=2).value


def get_cloumn_idx(file_path, sheet_name, cloumn_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(1, ws.max_column+1):
        if ws.cell(row=1, column=i).value == cloumn_name:
            break
    wb.close()
    return i

def get_case_value(file_path, sheet_name, case_name, cloumn_name):
    idx = get_cloumn_idx(file_path, sheet_name, cloumn_name)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == case_name:
            break
    wb.close()
    return ws.cell(row=i, column=idx).value

def get_case_list(file_path, sheet_name, case_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == case_name:
            break
    lists = []
    for j in range(2, ws.max_column+1):
        lists.append(ws.cell(row=i, column=j).value)
    wb.close()
    return lists

path = 'C:\\wq\\pyworkspace\\xlinkweb\\test_data\\login.xlsx'
s_name = 'locator'
l_name = 'password_loc'
col_name = 'value'
# val = get_loc_value(path, s_name, l_name)
# print val

# val = get_case_value(path, s_name, l_name, col_name)
# print val

val = get_case_list(path, s_name, l_name)
print val[0]