# coding=utf-8
from datetime import date, timedelta, datetime

import requests
import openpyxl
import random
import json
import function as func
import traceback
import operator
global false, true
false = False
true = True

from deployment_test import ApiTest

object_dict = {}
body_json = {}
keys_lists = []
keys_types = []
input_value = []
non_empty = []
min_values = []
max_values = []
keys_len = []
id_str = None
start_row = None
end_row = None
post_method = None
token = None
a = None
corp_id = None
login_url = 'http://iotapitest.bgycc.com/v2/corp_auth'
south_api_url = 'http://iotapitest.bgycc.com/v2/service/iot/publish?sdk=1'
north_api_url1 = 'http://iotapitest.bgycc.com/v2/service/iot/get_data?service=service_name&object=object_name&id=id_num'
north_api_url2 = 'http://iotapitest.bgycc.com/v2/service/iot/query?service=service_name&object=object_name'
headers = {"Content-Type": "application/json", "Access-Token": "token_value", "Api-Version": "2"}



def get_access_token(url, login_body):
    func.log(u'开始获取token')
    try:
        r = requests.post(url=url, json=login_body, headers=headers)
        func.log('r.text:' + r.text)
        res = eval(r.text)
        func.log('token:' + res['access_token'])
        modify_dict(headers, "Access-Token", res['access_token'])
        return res['corp_id']
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise


def get_random(type, min, max):
    if type == 'int32':
        if min is not None and max is not None:
            return random.randint(int(min), int(max))
        else:
            return random.randint(1, 100)
    elif type == 'float':
        if min is not None and max is not None:
            a = random.uniform(int(min), int(max))
            return round(a, 2)
        else:
            a = random.uniform(1, 2)
            return round(a, 2)
    elif type == 'boolean':
        return True


def get_invalid_value(type):
    if type == 'int32':
        return 1.1
    elif type == 'float':
        return 'a'
    elif type == 'boolean':
        return 'True'


def replace_url(url_base, dicts):
    for key in dicts.keys():
        url_base = url_base.replace(key, dicts[key])
    return url_base


def modify_dict(str_dict, key, value):
    str_dict[key] = value
    return str_dict


def modify_body_base(dic):
    body_json['service_id'] = dic["service_name"]
    body_json['table'] = dic["object_name"]
    body_json['operation'] = post_method


def get_start_end_rows(dic, path):
    func.log(path)
    global start_row
    global end_row
    global post_method
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']
    for i in range(2, ws.max_row + 1):
        row = ws.cell(row=i, column=2).value
        if row == dic["object_name"]:
            start_row = i
            post_method = ws.cell(row=i, column=4).value
            break
    for j in range(start_row + 1, ws.max_row + 1):
        row = ws.cell(row=j, column=2).value
        if row != dic["object_name"]:
            if row is not None and j != ws.max_row:
                end_row = j - 1
                break
            if row is None and j != ws.max_row + 1:
                pass
            if row is None and j == ws.max_row:
                end_row = ws.max_row


# print (start_row)
# print (end_row)


def get_objects_values(start, end, path):
    end = end + 1
    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']
    del keys_lists[:]
    del keys_types[:]
    del non_empty[:]
    del min_values[:]
    del max_values[:]
    del keys_len[:]
    for i in range(start, end):
        keys = ws.cell(row=i, column=5).value
        keys_lists.append(keys)
        types = ws.cell(row=i, column=7).value
        keys_types.append(types)
        empty = ws.cell(row=i, column=10).value
        non_empty.append(empty)
        min = ws.cell(row=i, column=11).value
        min_values.append(min)
        max = ws.cell(row=i, column=12).value
        max_values.append(max)
        lens = ws.cell(row=i, column=13).value
        keys_len.append(lens)


def create_data_normal():
    global id_str
    body_data_json = {}
    for i in range(0, len(keys_lists)):
        if keys_types[i] == 'string':
            if keys_lists[i] == 'id':
                id_str = func.random_string()
                body_data_json[keys_lists[i]] = id_str
            else:
                body_data_json[keys_lists[i]] = keys_lists[i]
                input_value.append(keys_lists[i])
        else:
            val = get_random(keys_types[i], min_values[i], max_values[i])
            body_data_json[keys_lists[i]] = val
            input_value.append(val)
    return body_data_json


def create_data_invalid_range():
    global id_str
    body_data_json = {}
    data_invalid_range = []
    for i in range(0, len(keys_lists)):
        if keys_types[i] == 'string':
            if keys_lists[i] == 'id':
                id_str = func.random_string()
                body_data_json[keys_lists[i]] = str(id_str)
            else:
                body_data_json[keys_lists[i]] = keys_lists[i]
                input_value.append(keys_lists[i])
        elif keys_types[i] == 'int32':
            if min_values[i] is not None and max_values[i] is not None:
                data_invalid_range.append(
                    keys_lists[i] + ' type:' + keys_types[i] + ' min:' + str(min_values[i]) + ", max:" + str(
                        max_values[i]))
                body_data_json[keys_lists[i]] = int(max_values[i]) + 1
            else:
                body_data_json[keys_lists[i]] = 1
        elif keys_types[i] == 'float':
            if min_values[i] is not None and max_values[i] is not None:
                data_invalid_range.append(
                    keys_lists[i] + ' type:' + keys_types[i] + ' min:' + str(min_values[i]) + ", max:" + str(
                        max_values[i]))
                body_data_json[keys_lists[i]] = float(max_values[i]) + 1
            else:
                body_data_json[keys_lists[i]] = 1.1
        elif keys_types[i] == 'boolean':
            body_data_json[keys_lists[i]] = True
    return body_data_json, data_invalid_range


def create_data_invalid_value():
    global id_str
    body_data_json = {}
    data_invalid_value = []
    for i in range(0, len(keys_lists)):
        if keys_types[i] == 'string':
            if keys_lists[i] == 'id':
                id_str = func.random_string()
                body_data_json[keys_lists[i]] = str(id_str)
            else:
                body_data_json[keys_lists[i]] = keys_lists[i]
                input_value.append(keys_lists[i])
        else:
            if keys_lists[i] == 'id':
                id_str = func.random_string()
                body_data_json[keys_lists[i]] = str(id_str)
            else:
                val = get_invalid_value(keys_types[i])
                body_data_json[keys_lists[i]] = val
                input_value.append(val)
                data_invalid_value.append(keys_lists[i] + ' type:' + keys_types[i])
    return body_data_json, data_invalid_value


def create_data_non_empty():
    body_data_json = {}
    data_non_empty = []
    for i in range(0, len(keys_lists)):
        if non_empty[i] == u'是':
            data_non_empty.append(keys_lists[i])
    return body_data_json, data_non_empty


def get_query_json(dicts):
    url = replace_url(north_api_url2, dicts)
    func.log(u'北向all接口url：' + url)
    body = {"limit": 1000}
    res_list = {}
    try:
        r = requests.post(url=url, json=body, headers=headers)
        res = eval(r.text)
        func.log(json.dumps(res))
        lists = res["list"]
        count_id = res["count"]
        for num in range(0, len(lists)):
            if str(lists[num]["id"]) == id_str:
                func.log(u'id=：' + str(id_str) + u' body：' + str(lists[num]))
                res_list = lists[num]
                break
        func.log(u'已有记录数：' + str(res["count"]))
        return count_id, res_list
    except Exception, e:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    finally:
        pass


def post_api(url, body):
    func.log(u'开始掉南向接口')
    try:
        r = requests.post(url=url, json=body, headers=headers)
        res = eval(r.text)
        func.log(u'南向接口响应：' + json.dumps(res))
        return res
    except Exception, e:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    finally:
        pass


def get_data_api(dicts):
    url = replace_url(north_api_url1, dicts)
    func.log(u'北向id接口url：' + url)
    body = {}
    try:
        r = requests.get(url=url, json=body, headers=headers)
        res = eval(r.text)
        func.log(u'北向接口通过id查询返回：' + json.dumps(res))
        return res
    except Exception, e:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    finally:
        pass


def assert_status(value):
    func.log(u'断言南向接口请求返回')
    try:
        assert value == "ok"
        func.log("success")
    except Exception:
        func.log(u'断言失败：实际内容为' + str(value))
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    finally:
        pass


def assert_invalid_value(invalid_value_lists, res):
    func.log(u'断言南向接口非法数据请求返回')
    global a
    a = 0
    for i in range(0, len(invalid_value_lists)):
        try:
            assert invalid_value_lists[i] in str(res)
        except Exception:
            a = 1
            func.log(u'断言失败：' + str(invalid_value_lists[i]) + u' 不在返回结果中')
            continue
    if a == 0:
        func.log(u'非法数据请求返回断言成功')
    else:
        assert a == 0


def assert_body_data(value1, value2):
    func.log(u'断言北向接口请求返回的body')
    # value1 = json.dumps(value1)
    # value2 = json.dumps(value2)
    func.log('a :' + str(value1))
    func.log('b :' + str(value2))
    func.log(operator.eq(value1, value2))
    try:
        assert operator.eq(value1, value2)
        func.log("success")
    except Exception:
        func.log(u'断言失败：' + str(value2))
        func.log(u'预期返回：' + str(value1))
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    finally:
        pass


def get_help_app(host, port):
    false = False
    true = True
    url = host + port + '/v2/plugin/apps'
    r = requests.get(url=url, json={}, headers=headers)
    res = eval(r.text)
    lists = res["list"]
    for i in lists:
        if i["plugin"] == 'helpdesk':
            url1 = host + port + '/v2/plugin/app/' + str(i["id"])
            requests.delete(url=url1, json={}, headers=headers)


def get_apps_aligenie_id(url):
    func.log(u'开始获取天猫精灵appid')
    func.log('url：' + url)
    try:
        r = requests.get(url=url, json={}, headers=headers)
        res = eval(r.text)
        func.log(u'获取天猫精灵appid响应：' + json.dumps(res))
        list = res['list']
        for i in list:
            if i['plugin'] == 'aligenie':
                return i['id']
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise


def set_aligenie_config(url, client_id, client_Secret):
    func.log(u'开始配置天猫精灵服务')
    body = {
        "name": u"天猫精灵",
        "enable": true,
        "config": {
            "scope": "",
            "client_secret": client_Secret,
            "client_id": client_id,
            "redirect_url": ""
        }
    }
    func.log('url：' + url)
    func.log('body：' + json.dumps(body))
    try:
        r = requests.put(url=url, json=body, headers=headers)
        res = eval(r.text)
        func.log(u'配置天猫精灵服务响应：' + json.dumps(res))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise


def get_device_summary(host, port):
    filter_id = ''
    func.log(u'开始获取异常设备数')
    exception_url = host + port + '/v2/wide-devices'
    exception_body = {
        "offset": 0,
        "limit": 1,
        "query": {
            "is_exception": {
                "$eq": true
            }
        }
    }
    try:
        r = requests.post(url=exception_url, json=exception_body, headers=headers)
        res = eval(r.text)
        exception_count = int(res['count'])
        func.log(u'异常设备数为：' + str(exception_count))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    func.log(u'开始获取过滤器配置')
    filter_url = host + port + '/v2/devices/filter/configs'
    filter_body = {
        "offset": 0,
        "limit": 1000,
        "filter": [
            "id",
            "name"
        ]
    }
    try:
        r = requests.post(url=filter_url, json=filter_body, headers=headers)
        res = eval(r.text)
        filter_id_list = res['list']
        for i in filter_id_list:
            if i['name'] == u'全部':
                filter_id = i['id']
                break
        func.log(u'全部过滤器为：' + filter_id)
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    func.log(u'开始获取设备总数和在线设备数')
    summary_url = host + port + '/v2/statistics/devices/filter/' + filter_id + '/summary'
    summary_body = {
        "offset": 0,
        "limit": 1000,
        "filter": [
            "id",
            "name"
        ]
    }
    try:
        r = requests.get(url=summary_url, json=summary_body, headers=headers)
        res = eval(r.text)
        total_devices = int(res['total'])
        online_devices = int(res['online'])
        offline_devices = total_devices - online_devices
        func.log(u'设备总数为：' + str(total_devices) + u' 在线数为：' + str(online_devices) + u' 离线数为：' + str(offline_devices))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    return total_devices, online_devices, offline_devices, exception_count


def get_alarm_summary(host, port):
    func.log(u'开始获取设备报警统计')
    week_ago = (date.today() + timedelta(days=-7)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S.999Z")
    summary_url = host + port + '/v2/statistics/alarm-state'
    summary_body = {
        "start_time": {
            "$gt": week_ago,
            "$lt": now
        }
    }
    try:
        r = requests.post(url=summary_url, json=summary_body, headers=headers)
        res = eval(r.text)
        alarms = int(res['alarms'])
        exceptions = int(res['exceptions'])
        today_alarms = int(res['today_alarms'])
        condition_alarms = int(res['condition_alarams'])
        func.log(u'正在报警为：' + str(alarms) + u' 今日异常设备总数为：' + str(exceptions) + u' 今日报警总数为：' + str(
            today_alarms) + u' 近7日报警总数为：' + str(condition_alarms))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    return alarms, exceptions, today_alarms, condition_alarms


def get_map_summary(host, port):
    func.log(u'开始获取地图设备总数')
    total_url = host + port + '/v2/wide-devices'
    total_body = {"offset": 0, "limit": 1, "query": {}}
    try:
        r = requests.post(url=total_url, json=total_body, headers=headers)
        res = eval(r.text)
        total = int(res['count'])
        func.log(u'地图设备总数为：' + str(total))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    func.log(u'开始获取地图在线设备数')
    online_url = host + port + '/v2/wide-devices'
    online_body = {"offset": 0, "limit": 1, "query": {"is_online": {"$eq": true}}}
    try:
        r = requests.post(url=online_url, json=online_body, headers=headers)
        res = eval(r.text)
        online = int(res['count'])
        func.log(u'地图在线设备数为：' + str(online))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    func.log(u'开始获取地图异常设备数')
    exception_url = host + port + '/v2/wide-devices'
    exception_body = {"offset": 0, "limit": 1, "query": {"is_exception": {"$eq": true}}}
    try:
        r = requests.post(url=exception_url, json=exception_body, headers=headers)
        res = eval(r.text)
        exception = int(res['count'])
        func.log(u'地图异常设备数为：' + str(exception))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise
    return total, online, exception


def set_corp_setting(host, port, setting):
    """设置是否激活验证"""
    corp_setting_url = host + port + '/v2/corp_setting'
    func.log(corp_setting_url)
    corp_setting_body = {"is_user_active":setting}
    func.log(u'开始设置是否激活验证为：'+str(setting))
    try:
        r = requests.post(url=corp_setting_url, json=corp_setting_body, headers=headers)
        res = eval(r.text)
        result = res['is_user_active']
        func.log(u'是否激活验证已设置为为：' + str(result))
    except Exception:
        func.log(('traceback.format_exc():\n%s' % traceback.format_exc()))
        raise

