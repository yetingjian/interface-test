# encoding: utf-8
import openpyxl


def get_locator_value(file_path, sheet_name, loc_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=1).value == loc_name:
            break
    wb.close()
    return ws.cell(row=i, column=2).value


def get_case_value(file_path, sheet_name, case_name, column_name):
    idx = get_column_idx(file_path, sheet_name, column_name)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == case_name:
            break
    wb.close()
    return ws.cell(row=i, column=idx).value


def get_column_idx(file_path, sheet_name, column_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(1, ws.max_column+1):
        if ws.cell(row=1, column=i).value == column_name:
            break
    wb.close()
    return i


def get_case_list(file_path, sheet_name, case_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    lists = []
    a = 0
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == case_name:
            break
    else:
        a = 1
    if a == 0:
        for j in range(3, ws.max_column+1):
            lists.append(ws.cell(row=i, column=j).value)
        wb.close()
        return lists
    else:
        return lists


def get_case_status(file_path, sheet_name, case_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == case_name:
            return ws.cell(row=i, column=2).value
