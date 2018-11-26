# encoding: utf-8

import os
import requests
import openpyxl
import traceback
global false, true
false = False
true = True

def find_path():
    base_dir = os.getcwd()
    base_dir = base_dir.replace('\\', '/')
    return base_dir


def load_data():
    file_path = find_path() + '/data.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet1']
    lists = []
    for i in range(2, ws.max_row+1):
        row_lists = []
        for j in range(1, ws.max_column+1):
            value = ws.cell(row=i, column=j).value
            if value is None:
                value = ''
            row_lists.append(value)
        lists.append(row_lists)
    return lists


def modify_dict(str_dict, key, value):
    str_dict[key] = value
    return str_dict


def get_access_token():
    try:
        r = requests.post(url=login_url, json=login_body, headers=headers)
        res = eval(r.text)
        modify_dict(headers, "Access-Token", res['access_token'])
    except Exception:
        print('traceback.format_exc():\n%s' % traceback.format_exc())
        raise


def add_data_points():
    source = 0
    type = 0
    read = 0
    url = host + port + '/v2/product/'+value_list[0][0]+'/datapoint'
    for i in range(len(value_list)):
        if value_list[i][5] == u'应用设置':
            source = 1
        elif value_list[i][5] == u'设备上报':
            source = 3
        if value_list[i][6] == u'布尔':
            type = 1
        elif value_list[i][6] == u'单字节':
            type = 2
        elif value_list[i][6] == u'int16有符号':
            type = 3
        elif value_list[i][6] == u'int32有符号':
            type = 4
        elif value_list[i][6] == u'浮点':
            type = 5
        elif value_list[i][6] == u'字符串':
            type = 6
        elif value_list[i][6] == u'int16无符号':
            type = 8
        elif value_list[i][6] == u'int32无符号':
            type = 9
        if value_list[i][10] == u'可读写':
            read = 1
        elif value_list[i][10] == u'只读':
            read = 0
        body = {
            "name": value_list[i][3],
            "field_name": value_list[i][4],
            "type": type,
            "index": value_list[i][1],
            "description": value_list[i][11],
            "symbol": value_list[i][9],
            "source": source,
            "is_read": true,
            "is_write": read,
            "min": value_list[i][7],
            "max": value_list[i][8]
        }
        print body
        try:
            requests.post(url=url, json=body, headers=headers)
            # res = eval(r.text)
        except Exception:
            print('traceback.format_exc():\n%s' % traceback.format_exc())
            raise


def get_all_data_points():
    url = host + port + '/v2/product/' + value_list[0][0] + '/datapoints'
    try:
        r = requests.get(url=url, json={}, headers=headers)
        res = eval(r.text)
        return res
    except Exception:
        print('traceback.format_exc():\n%s' % traceback.format_exc())
        raise


def get_all_point_id(res):
    ids = []
    for i in range(len(res)):
        ids.append(res[i]['id'])
    return ids


def delete_all_data_point():
    rest = get_all_data_points()
    id_list = get_all_point_id(rest)
    for i in id_list:
        url = host + port + '/v2/product/' + pid + '/datapoint/' + i
        try:
            requests.delete(url=url, json={}, headers=headers)
        except Exception:
            print('traceback.format_exc():\n%s' % traceback.format_exc())
            raise


if __name__ == '__main__':
    host = 'http://dev-api.xlink.cn'            # API地址
    port = ''                                   # API端口
    pid = '1607d2b6900000011607d2b69000ba01'    # 删除数据端点时才需配置
    login_url = host + port + '/v2/corp_auth'
    login_body = {"account": "gyb@xlink.cn", "password": "Test1234"}                # 登录账号，自行修改
    headers = {"Content-Type": "application/json", "Access-Token": "token_value"}
    value_list = load_data()  # 读取数据端点数据表
    get_access_token()        # 获取token
    add_data_points()       # 添加数据端点
    # delete_all_data_point()   # 删除所有数据端点


